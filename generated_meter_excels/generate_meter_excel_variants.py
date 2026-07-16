from pathlib import Path
import math
import random
import sys
from copy import copy

from openpyxl import load_workbook

ROOT = Path(r"D:\WorkSpace\NQI_Project_Server")
sys.path.insert(0, str(ROOT))
from app.data_processing import parse_excel_file

SOURCE = Path(r"D:\BaiduSyncdisk\NQI\数据集\电量数据 - 数值\TA3310三相表数据.xlsx")
OUT_DIR = ROOT / "generated_meter_excels"
OUT_DIR.mkdir(parents=True, exist_ok=True)

# 上位机固定条形图 x 轴来自每个指标块的 C 列相角和 A 列里的电流，生成时保持这些单元格不变。
METRIC_BLOCKS = {
    "功率W": (6, 41),
    "电压": (48, 83),
    "电流": (89, 124),
    "相角": (132, 167),
}
PHASE_BLOCKS = [(5, 6, 7, 8), (9, 10, 11, 12), (13, 14, 15, 16)]  # E:F/G:H, I:J/K:L, M:N/O:P
SHEET_FACTOR = {"A": 1.000, "B": 0.985, "C": 1.012}
PHASE_FACTOR = {0: 1.00, 1: 0.999, 2: 1.001}

VARIANTS = [
    {
        "file": "NQI_meter_01_normal.xlsx",
        "desc": "正常波动：整体误差小，适合验证正常显示。",
        "kind": "normal",
        "seed": 101,
        "dev_base": 0.00025,
    },
    {
        "file": "NQI_meter_02_stable.xlsx",
        "desc": "低误差稳定：两台设备几乎一致，误差接近 0。",
        "kind": "stable",
        "seed": 202,
        "dev_base": 0.00005,
    },
    {
        "file": "NQI_meter_03_high_power.xlsx",
        "desc": "功率偏高：功率块整体抬高，部分点明显变大。",
        "kind": "high_power",
        "seed": 303,
        "dev_base": 0.0008,
    },
    {
        "file": "NQI_meter_04_voltage_fault.xlsx",
        "desc": "电压异常：电压块存在欠压/过压波动。",
        "kind": "voltage_fault",
        "seed": 404,
        "dev_base": 0.0012,
    },
    {
        "file": "NQI_meter_05_current_spike.xlsx",
        "desc": "电流突增：电流块在高电流测试点出现尖峰。",
        "kind": "current_spike",
        "seed": 505,
        "dev_base": 0.0010,
    },
    {
        "file": "NQI_meter_06_mixed_fault.xlsx",
        "desc": "相角漂移混合异常：相角块偏移，同时局部功率/电压有异常点。",
        "kind": "mixed_fault",
        "seed": 606,
        "dev_base": 0.0018,
    },
]


def _current_from_range(range_text: str) -> float:
    text = str(range_text or "")
    if "," not in text:
        return 0.0
    raw = text.split(",", 1)[1].strip()
    digits = "".join(ch for ch in raw if ch.isdigit() or ch == ".")
    value = float(digits or 0)
    return value / 1000.0 if "mA" in raw else value


def _base_value(metric: str, angle: float, current_a: float, sheet_name: str, phase_index: int, point_index: int, variant: dict) -> float:
    sf = SHEET_FACTOR.get(sheet_name, 1.0)
    pf = PHASE_FACTOR.get(phase_index, 1.0)
    kind = variant["kind"]
    angle_factor = 1.0 if abs(angle) < 1e-9 else (0.5 if abs(angle - 60) < 1e-9 else 0.5)

    if metric == "功率W":
        base = 240.0 * current_a * angle_factor
        if kind == "high_power":
            base *= 1.45 + 0.08 * math.sin(point_index / 3)
        elif kind == "mixed_fault" and point_index in (8, 17, 26, 35):
            base *= 1.85
        elif kind == "stable":
            base *= 0.995
    elif metric == "电压":
        base = 480.0 if current_a <= 0.2 else 240.0
        if kind == "voltage_fault":
            if point_index % 6 in (0, 1):
                base *= 0.88
            elif point_index % 6 == 5:
                base *= 1.12
        elif kind == "mixed_fault" and point_index in (4, 13, 22, 31):
            base *= 0.82
        elif kind == "stable":
            base *= 1.0002
    elif metric == "电流":
        base = current_a * 2.0
        if kind == "current_spike" and current_a >= 5:
            base *= 1.35 + 0.15 * (point_index % 3)
        elif kind == "mixed_fault" and current_a >= 10 and phase_index == 1:
            base *= 1.6
        elif kind == "stable":
            base *= 0.998
    else:  # 相角
        base = angle * 2.0 if angle else 0.0
        if kind == "phase_drift" or kind == "mixed_fault":
            base += (phase_index + 1) * 8 + (point_index % 4) * 3
        elif kind == "stable":
            base += 0.2 * phase_index
    return base * sf * pf


def _write_variant(variant: dict) -> Path:
    random.seed(variant["seed"])
    wb = load_workbook(SOURCE)
    for ws in wb.worksheets:
        sheet_name = ws.title
        # 说明写到空白区，不影响解析逻辑，便于人工识别样例用途。
        ws["A1"] = variant["desc"]
        for metric, (start_row, end_row) in METRIC_BLOCKS.items():
            for row in range(start_row, end_row + 1):
                point_index = row - start_row
                angle = float(ws.cell(row=row, column=3).value or 0)
                current_a = _current_from_range(ws.cell(row=row, column=1).value)
                for phase_index, (ref_col, cmp_col, err_col, ppm_col) in enumerate(PHASE_BLOCKS):
                    base = _base_value(metric, angle, current_a, sheet_name, phase_index, point_index, variant)
                    noise = 1 + random.uniform(-0.0015, 0.0015)
                    reference = base * noise
                    direction = -1 if (point_index + phase_index + variant["seed"]) % 2 else 1
                    deviation = variant["dev_base"] * direction * (1 + (point_index % 5) * 0.35)
                    if variant["kind"] in ("high_power", "voltage_fault", "current_spike", "mixed_fault") and point_index in (5, 11, 23, 34):
                        deviation *= 8
                    compared = reference * (1 - deviation)

                    ws.cell(row=row, column=ref_col).value = round(reference, 6)
                    ws.cell(row=row, column=cmp_col).value = round(compared, 6)
                    error_percent = (reference - compared) / compared if compared else 0.0
                    ws.cell(row=row, column=err_col).value = round(error_percent, 12)
                    ws.cell(row=row, column=ppm_col).value = round(error_percent * 1000000, 6)
    output = OUT_DIR / variant["file"]
    wb.save(output)
    return output


outputs = []
for variant in VARIANTS:
    path = _write_variant(variant)
    payload = parse_excel_file(path)
    outputs.append({
        "file": str(path),
        "sheet_count": payload.get("sheet_count"),
        "chart_point_count": payload.get("chart_point_count"),
        "chart_value_count": payload.get("chart_value_count"),
        "error_value_count": payload.get("error_value_count"),
        "max_numeric_value": payload.get("max_numeric_value"),
        "min_numeric_value": payload.get("min_numeric_value"),
    })

for item in outputs:
    print(item)




