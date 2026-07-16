from pathlib import Path
import json
import sys

ROOT = Path(r"D:\WorkSpace\NQI_Project_Server")
sys.path.insert(0, str(ROOT))

from openpyxl import load_workbook
from app.data_processing import parse_excel_file

sample = Path(r"D:\BaiduSyncdisk\NQI\数据集\电量数据 - 数值\TA3310三相表数据.xlsx")
wb = load_workbook(sample, data_only=False)
print("sheets", wb.sheetnames)
for ws in wb.worksheets:
    print("sheet", ws.title, "rows", ws.max_row, "cols", ws.max_column)
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 14), values_only=True):
        print(row[:10])
    print("---")

payload = parse_excel_file(sample)
summary = {
    "sheet_count": payload.get("sheet_count"),
    "sheet_names": payload.get("summary", {}).get("sheet_names"),
    "chart_point_count": payload.get("chart_point_count"),
    "chart_value_count": payload.get("chart_value_count"),
    "error_value_count": payload.get("error_value_count"),
}
print("parse_summary", json.dumps(summary, ensure_ascii=False, indent=2))
first_sheet = next(iter(payload.get("parsed_data", {}).values()))
for metric in first_sheet.get("data", []):
    x = metric.get("data", {}).get("x", {})
    print("metric", metric.get("name"), "x_name", x.get("name"), "x_count", len(x.get("data", [])), "x_preview", x.get("data", [])[:8])
