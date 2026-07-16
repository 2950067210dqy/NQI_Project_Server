from pathlib import Path
from openpyxl import load_workbook
sample = Path(r"D:\BaiduSyncdisk\NQI\数据集\电量数据 - 数值\TA3310三相表数据.xlsx")
wb = load_workbook(sample)
for ws in wb.worksheets:
    print(ws.title, list(ws.merged_cells.ranges)[:50], 'count=', len(ws.merged_cells.ranges))
    for row in (6,42,78,114):
        print('row', row, [type(ws.cell(row=row, column=c)).__name__ for c in range(1,17)])
        print([ws.cell(row=row, column=c).value for c in range(1,17)])
